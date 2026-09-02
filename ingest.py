# ingest.py
import os
import json
import time
from typing import List, Dict, Any
from tqdm import tqdm

from langchain_core.documents import Document

# Document loaders (community package)
try:
    from langchain_community.document_loaders import PyPDFLoader, TextLoader, Docx2txtLoader
except Exception as exc:
    raise ImportError(
        "Please install the document-loading stack: python -m pip install langchain-community pypdf python-docx docx2txt"
    ) from exc

# Text splitter (splitters package)
try:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
except Exception as exc:
    raise ImportError(
        "Please install langchain-text-splitters: python -m pip install langchain-text-splitters"
    ) from exc

# Vector store and embeddings (langchain)
try:
    from langchain_chroma import Chroma
except Exception:
    try:
        from langchain_community.vectorstores import Chroma
    except Exception:
        try:
            from langchain.vectorstores import Chroma
        except Exception as exc:
            raise ImportError(
                "Please install langchain and related packages: python -m pip install langchain langchain-community langchain-text-splitters chromadb sentence-transformers langchain-openai langchain-chroma langchain-huggingface"
            ) from exc

try:
    from langchain_huggingface import HuggingFaceEmbeddings
except Exception:
    try:
        from langchain_community.embeddings import HuggingFaceEmbeddings
    except Exception:
        try:
            from langchain.embeddings import HuggingFaceEmbeddings
        except Exception as exc:
            raise ImportError(
                "Please install langchain-community or langchain with HuggingFace embeddings support."
            ) from exc

try:
    from langchain_openai import OpenAIEmbeddings
except Exception:
    try:
        from langchain_community.embeddings import OpenAIEmbeddings
    except Exception:
        try:
            from langchain.embeddings import OpenAIEmbeddings
        except Exception:
            OpenAIEmbeddings = None

try:
    from langchain_google_genai import GoogleGenerativeAIEmbeddings
except Exception:
    try:
        from langchain_community.embeddings import GoogleGenerativeAIEmbeddings
    except Exception:
        try:
            from langchain.embeddings import GoogleGenerativeAIEmbeddings
        except Exception:
            GoogleGenerativeAIEmbeddings = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from dotenv import load_dotenv

# ---------- CONFIG ----------
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
CODE_DIR = os.path.join(PROJECT_ROOT, "Code")
DATA_DIR = os.path.join(PROJECT_ROOT, "Data")
CHROMA_PERSIST_DIR = os.path.join(PROJECT_ROOT, "chroma_db")
MANIFEST_FILE = os.path.join(PROJECT_ROOT, "ingest_manifest.json")  # file-level manifest
CHUNK_SIZE = 1000
CHUNK_OVERLAP = 150
HF_MODEL_NAME = "sentence-transformers/all-MiniLM-L6-v2"
OPENAI_EMBEDDING_MODEL = "text-embedding-3-small"
# ----------------------------

# Load .env from Code folder if present
env_path = os.path.join(CODE_DIR, ".env")
if os.path.exists(env_path):
    load_dotenv(env_path)
else:
    load_dotenv()

def read_config() -> Dict[str, Any]:
    cfg_path = os.path.join(CODE_DIR, "config.json")
    if not os.path.exists(cfg_path):
        return {}
    with open(cfg_path, "r", encoding="utf8") as f:
        return json.load(f)

def list_files(folder: str) -> List[str]:
    exts = (".pdf", ".txt", ".docx", ".xlsx", ".xls")
    return [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(exts)]


def _load_excel_documents(path: str, source_name: str) -> List[Dict[str, Any]]:
    """Extract worksheet rows from Excel files into document chunks."""
    docs: List[Dict[str, Any]] = []
    if load_workbook is None:
        print(f"Skipping Excel file {path}: openpyxl is not installed.")
        return docs

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                cleaned = ["" if cell is None else str(cell).strip() for cell in row]
                text = " | ".join(cleaned).strip()
                if not text:
                    continue
                docs.append({
                    "page_content": f"Sheet: {sheet} | Row {row_idx + 1} | {text}",
                    "metadata": {"source": source_name, "sheet_name": sheet, "row_number": row_idx + 1},
                })
        workbook.close()
    except Exception as e:
        print(f"Error loading Excel file {path}: {e}")
    return docs


def _extract_text_and_meta_from_loader_output(item: Any, source_name: str) -> Dict[str, Any]:
    """
    Normalize loader output to a simple dict with 'page_content' and 'metadata'.
    Works whether loader returns a langchain Document-like object or plain strings.
    """
    metadata = {}
    # try common attributes
    text = None
    if hasattr(item, "page_content"):
        text = item.page_content
    elif hasattr(item, "text"):
        text = item.text
    elif isinstance(item, str):
        text = item
    else:
        # fallback to str()
        text = str(item)

    # metadata extraction
    if not isinstance(item, str) and hasattr(item, "metadata"):
        try:
            metadata = dict(item.metadata)
        except Exception:
            metadata = {}
    # ensure source and page_number if available
    metadata.setdefault("source", source_name)
    if "page" in metadata and "page_number" not in metadata:
        metadata["page_number"] = metadata.get("page")
    return {"page_content": text, "metadata": metadata}

def load_documents_for_file(path: str) -> List[Dict[str, Any]]:
    """Load a single file and return a list of normalized dicts: {'page_content','metadata'}."""
    basename = os.path.basename(path)
    docs: List[Dict[str, Any]] = []
    try:
        if path.lower().endswith(".pdf"):
            loader = PyPDFLoader(path)
            loaded = loader.load_and_split()
            for item in loaded:
                docs.append(_extract_text_and_meta_from_loader_output(item, basename))
        elif path.lower().endswith(".txt"):
            loader = TextLoader(path, encoding="utf8")
            loaded = loader.load()
            for item in loaded:
                docs.append(_extract_text_and_meta_from_loader_output(item, basename))
        elif path.lower().endswith(".docx"):
            loader = Docx2txtLoader(path)
            loaded = loader.load()
            for item in loaded:
                docs.append(_extract_text_and_meta_from_loader_output(item, basename))
        elif path.lower().endswith((".xlsx", ".xls")):
            docs.extend(_load_excel_documents(path, basename))
        else:
            print(f"Skipping unsupported file: {path}")
    except Exception as e:
        print(f"Error loading {path}: {e}")
    return docs

def chunk_documents(docs: List[Dict[str, Any]], chunk_size: int, chunk_overlap: int) -> List[Dict[str, Any]]:
    """
    Split each document's text into chunks while preserving metadata.
    Uses RecursiveCharacterTextSplitter.split_text to avoid requiring langchain Document class.
    """
    splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=chunk_overlap)
    chunks: List[Dict[str, Any]] = []
    for doc_idx, d in enumerate(docs):
        text = d.get("page_content", "")
        meta = dict(d.get("metadata", {}))
        # prefer splitter.split_text if available, else fallback to split_documents API
        if hasattr(splitter, "split_text"):
            parts = splitter.split_text(text)
        else:
            # fallback: try split_documents with a valid LangChain Document
            try:
                tmp_docs = splitter.split_documents([Document(page_content=text, metadata={})])
                parts = [t.page_content for t in tmp_docs]
            except Exception:
                # last resort: naive chunking
                parts = [text[i:i+chunk_size] for i in range(0, max(1, len(text)), chunk_size - chunk_overlap)]
        for i, part in enumerate(parts):
            chunk_meta = dict(meta)
            chunk_meta.setdefault("source", meta.get("source", "unknown"))
            chunk_meta.setdefault("chunk_id", f"{chunk_meta.get('source')}_doc{doc_idx}_chunk{i}_{int(time.time())}")
            chunks.append({"page_content": part, "metadata": chunk_meta})
    return chunks

def choose_embeddings(config: dict):
    configured_provider = str(config.get("provider", "")).lower()
    openai_key = os.getenv("OPENAI_API_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")

    preferred = configured_provider if configured_provider in {"openai", "gemini", "local", "huggingface"} else ""

    if preferred in {"local", "huggingface"}:
        print(f"Using HuggingFace embeddings ({HF_MODEL_NAME}) — configured local mode.")
        try:
            return HuggingFaceEmbeddings(model_name=HF_MODEL_NAME)
        except Exception as exc:
            raise RuntimeError(
                "Unable to initialize the local HuggingFace embedding model. Check internet access or rerun after the model downloads successfully."
            ) from exc

    provider = preferred if preferred in {"openai", "gemini"} else ""
    if not provider:
        provider = "openai" if openai_key else "gemini" if gemini_key else ""
    elif provider == "gemini" and not gemini_key and openai_key:
        provider = "openai"
    elif provider == "openai" and not openai_key and gemini_key:
        provider = "gemini"

    if provider == "openai" and openai_key and OpenAIEmbeddings is not None:
        print("Using OpenAI embeddings (OPENAI_API_KEY detected).")
        try:
            return OpenAIEmbeddings(model=OPENAI_EMBEDDING_MODEL)
        except Exception:
            print("OpenAI embeddings failed; falling back to local HuggingFace embeddings.")

    if provider == "gemini" and gemini_key and GoogleGenerativeAIEmbeddings is not None:
        print("Using Gemini embeddings (GEMINI_API_KEY/GOOGLE_API_KEY detected).")
        try:
            return GoogleGenerativeAIEmbeddings(model="models/text-embedding-004")
        except Exception:
            print("Gemini embeddings failed; falling back to local HuggingFace embeddings.")

    if openai_key and OpenAIEmbeddings is not None:
        print("Using OpenAI embeddings (OPENAI_API_KEY detected).")
        try:
            return OpenAIEmbeddings(model=OPENAI_EMBEDDING_MODEL)
        except Exception:
            print("OpenAI embeddings failed; falling back to local HuggingFace embeddings.")

    if gemini_key and GoogleGenerativeAIEmbeddings is not None:
        print("Using Gemini embeddings (GEMINI_API_KEY/GOOGLE_API_KEY detected).")
        try:
            return GoogleGenerativeAIEmbeddings(model="models/text-embedding-004")
        except Exception:
            print("Gemini embeddings failed; falling back to local HuggingFace embeddings.")

    print(f"Using HuggingFace embeddings ({HF_MODEL_NAME}) — local fallback.")
    try:
        return HuggingFaceEmbeddings(model_name=HF_MODEL_NAME)
    except Exception as exc:
        raise RuntimeError(
            "Unable to initialize the local HuggingFace embedding model. Check internet access or rerun after the model downloads successfully."
        ) from exc

def load_manifest(manifest_path: str) -> Dict[str, float]:
    if not os.path.exists(manifest_path):
        return {}
    try:
        with open(manifest_path, "r", encoding="utf8") as f:
            data = json.load(f)
            return {k: float(v) for k, v in data.items()}
    except Exception:
        return {}

def save_manifest(manifest: Dict[str, float], manifest_path: str):
    with open(manifest_path, "w", encoding="utf8") as f:
        json.dump(manifest, f, indent=2)

def _persist_if_supported(vectordb) -> None:
    persist = getattr(vectordb, "persist", None)
    if callable(persist):
        persist()

def persist_chunks_upsert(chunks: List[Dict[str, Any]], persist_directory: str, embeddings):
    """Upsert chunks into an existing Chroma DB (create DB if missing)."""
    texts = [c["page_content"] for c in chunks]
    metadatas = [c["metadata"] for c in chunks]

    if os.path.exists(persist_directory) and os.listdir(persist_directory):
        vectordb = Chroma(persist_directory=persist_directory, embedding_function=embeddings)
        try:
            vectordb.add_texts(texts, metadatas=metadatas)
            _persist_if_supported(vectordb)
        except Exception as e:
            print("Warning: add_texts failed, recreating DB. Error:", e)
            vectordb = Chroma.from_texts(texts, embeddings, metadatas=metadatas, persist_directory=persist_directory)
            _persist_if_supported(vectordb)
    else:
        vectordb = Chroma.from_texts(texts, embeddings, metadatas=metadatas, persist_directory=persist_directory)
        _persist_if_supported(vectordb)
    return vectordb

def run_ingest(upsert: bool = True):
    config = read_config()
    files = list_files(DATA_DIR)
    if not files:
        raise SystemExit(f"No documents found in {DATA_DIR}. Add PDFs/TXT/DOCX and retry.")

    manifest = load_manifest(MANIFEST_FILE)
    to_process = []
    new_manifest = manifest.copy()

    for f in files:
        mtime = os.path.getmtime(f)
        basename = os.path.basename(f)
        prev = manifest.get(basename)
        if prev is None or float(prev) < float(mtime):
            to_process.append(f)
            new_manifest[basename] = mtime

    if not to_process:
        print("No new or changed files to ingest.")
        embeddings = choose_embeddings(config)
        if os.path.exists(CHROMA_PERSIST_DIR) and os.listdir(CHROMA_PERSIST_DIR):
            return Chroma(persist_directory=CHROMA_PERSIST_DIR, embedding_function=embeddings)
        else:
            raise SystemExit("No existing Chroma DB found. Add files to Data/ and re-run.")

    print(f"Processing {len(to_process)} file(s):")
    all_chunks: List[Dict[str, Any]] = []
    for path in tqdm(to_process, desc="Loading files"):
        docs = load_documents_for_file(path)
        if not docs:
            print(f"No text extracted from {path}. It may be scanned; consider OCR.")
            continue
        chunks = chunk_documents(docs, CHUNK_SIZE, CHUNK_OVERLAP)
        all_chunks.extend(chunks)

    if not all_chunks:
        raise SystemExit("No chunks created from the provided files.")

    print(f"Created {len(all_chunks)} chunks. Generating embeddings and persisting to Chroma...")
    embeddings = choose_embeddings(config)
    vectordb = persist_chunks_upsert(all_chunks, CHROMA_PERSIST_DIR, embeddings)

    save_manifest(new_manifest, MANIFEST_FILE)
    print("Ingestion complete.")
    return vectordb

if __name__ == "__main__":
    db = run_ingest(upsert=True)

    # --- quick smoke test (prints top 3 matches for a sample query) ---
    try:
        emb = choose_embeddings(read_config())
        db = Chroma(persist_directory=CHROMA_PERSIST_DIR, embedding_function=emb)
        results = db.similarity_search("What is the leave carry forward rule?", k=3)
        print("\n--- Smoke test results ---")
        for r in results:
            meta = getattr(r, "metadata", None)
            if not isinstance(meta, dict):
                meta = {}
            content = getattr(r, "page_content", None)
            if content is None and isinstance(r, dict):
                content = r.get("page_content", str(r))
            if content is None:
                content = str(r)
            print("SOURCE:", meta.get("source"), "PAGE:", meta.get("page_number"))
            print(content[:300].replace("\n", " "), "...\n")
    except Exception as e:
        print("Smoke test skipped (embedding or retrieval error):", e)

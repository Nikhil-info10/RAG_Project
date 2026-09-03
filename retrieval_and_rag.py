# retrieval_and_rag.py
import os
import json
import re
import requests
from typing import List, Dict, Any
from textwrap import shorten

from langchain_core.documents import Document

try:
    from langchain_chroma import Chroma
    from langchain_huggingface import HuggingFaceEmbeddings
except ImportError:
    raise ImportError("Please install langchain-chroma and langchain-huggingface packages.")

try:
    from langchain_openai import OpenAI
except ImportError:
    OpenAI = None

try:
    from langchain_community.retrievers import BM25Retriever
except ImportError:
    raise ImportError("Please install langchain-community and rank_bm25 for BM25Retriever.")

from sentence_transformers import CrossEncoder  # for reranking

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from ddgs import DDGS
except ImportError:
    DDGS = None

CHROMA_DIR = "chroma_db"
DATA_DIR = "Data"
HF_MODEL_NAME = "sentence-transformers/all-MiniLM-L6-v2"
PROVENANCE_FILE = "chroma_provenance.json"
MAX_CONTEXT_CHARS = 4000
MAX_HISTORY_TURNS = 6
NOT_FOUND_MESSAGE = "Information not found in the available documents."

BING_API_KEY = os.getenv("BING_API_KEY")
BING_ENDPOINT = "https://api.bing.microsoft.com/v7.0/search"
POLICY_TERMS = (
    "absent",
    "violation",
    "misconduct",
    "rule",
    "policy",
    "terminate",
    "termination",
    "hit someone",
    "hit an employee",
    "fight",
    "assault",
    "violence",
    "physical altercation",
)
LEAVE_TERMS = (
    "leave",
    "vacation",
    "pto",
    "sick",
    "personal",
    "annual",
    "days off",
    "time off",
    "absent",
)
EMPLOYEE_TERMS = (
    "employee",
    "staff",
    "work at",
    "work in",
    "works at",
    "works in",
    "employed by",
    "employee of",
    "employment status",
    "company employee",
    "organization employee",
    "department",
    "job title",
    "role",
    "position",
    "salary",
    "age",
    "location",
    "joined",
    "joining",
    "performance",
    "rating",
)
NON_EMPLOYEE_TERMS = (
    "mobile",
    "phone",
    "iphone",
    "buy",
    "purchase",
    "product",
    "weather",
    "news",
)

reranker = CrossEncoder("cross-encoder/ms-marco-MiniLM-L-6-v2")

def external_search(query: str) -> str:
    if BING_API_KEY:
        headers = {"Ocp-Apim-Subscription-Key": BING_API_KEY}
        params = {"q": query, "count": 3}
        try:
            resp = requests.get(BING_ENDPOINT, headers=headers, params=params, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            snippets = [w.get("snippet", "") for w in data.get("webPages", {}).get("value", [])]
            if snippets:
                return "\n".join(snippets)
        except Exception:
            pass

    if DDGS is not None:
        try:
            results = []
            with DDGS() as search_client:
                for result in search_client.text(query, max_results=5):
                    title = result.get("title", "")
                    url = result.get("href", "")
                    body = result.get("body", "")
                    results.append(f"{title}\nURL: {url}\n{body}")
            if results:
                return "\n\n".join(results)
        except Exception:
            pass

    try:
        response = requests.get(
            "https://api.duckduckgo.com/",
            params={"q": query, "format": "json", "no_html": 1, "skip_disambig": 1},
            timeout=15,
        )
        response.raise_for_status()
        data = response.json()
        results = []
        if data.get("AbstractText"):
            results.append(data["AbstractText"])
        for topic in data.get("RelatedTopics", []):
            if topic.get("Text"):
                results.append(topic["Text"])
            if len(results) >= 5:
                break
        if results:
            return "\n".join(results)
        return "No web results found."
    except Exception as exc:
        return f"External search failed: {exc}"


def needs_external_search(answer: str) -> bool:
    answer_lower = answer.lower()
    return any(marker in answer_lower for marker in (
        "i don't know",
        "i do not know",
        "not contained in the context",
        "not available in the provided",
        "not found in the context",
        "not found in the provided",
        "information not found in the available documents",
        "information not found in web search results",
        "no llm executed",
        "llm call failed",
    ))


def normalize_not_found(answer: str) -> str:
    if needs_external_search(answer):
        return NOT_FOUND_MESSAGE
    return answer


def answer_from_external_search(question: str) -> str:
    web_context = external_search(question)
    if web_context.startswith("External search failed") or web_context == "No web results found.":
        return NOT_FOUND_MESSAGE
    prompt = f"""Answer the user's question using the web-search results below.
Do not claim information that is not supported by the results. If the results do not answer the question, reply exactly: "Information not found in web search results."

Web-search results:
{web_context}

User question:
{question}

Answer concisely and mention that the information came from web search.
"""
    answer = call_llm(prompt)
    if answer == "No LLM executed.":
        return "No LLM provider is configured. Add OPENAI_API_KEY to Code/.env."
    if answer.startswith("[LLM call failed:"):
        return answer
    return normalize_not_found(answer)

def load_provenance_with_fallback(path: str) -> Dict[str, Dict[str, Any]]:
    try:
        with open(path, "r", encoding="utf8") as f:
            prov = json.load(f)
        return {p["chunk_id"]: p for p in prov if p.get("chunk_id")}
    except Exception:
        return {}

def assemble_context(docs: List[Any], max_chars: int = MAX_CONTEXT_CHARS) -> str:
    parts = []
    for d in docs:
        # Document objects have .metadata and .page_content
        if hasattr(d, "metadata") and hasattr(d, "page_content"):
            meta = d.metadata or {}
            text = d.page_content or ""
        else:
            # fallback if it's a dict
            meta = d.get("metadata", {})
            text = d.get("page_content", "")

        citation = meta.get("source", "unknown")
        preview = shorten(text.replace("\n", " "), width=800, placeholder=" ...")
        parts.append(f"[{citation}] {preview}")

    context = "\n\n".join(parts)
    if len(context) > max_chars:
        return context[: max_chars - 200] + "\n\n... [context truncated] ..."
    return context


def document_metadata(doc: Any) -> Dict[str, Any]:
    if hasattr(doc, "metadata"):
        return doc.metadata or {}
    return doc.get("metadata", {})


def is_leave_question(question: str) -> bool:
    question_lower = question.lower()
    return any(term in question_lower for term in LEAVE_TERMS)


def is_employee_question(question: str) -> bool:
    question_lower = question.lower()
    return is_leave_question(question) or any(term in question_lower for term in EMPLOYEE_TERMS)


def source_names(docs: List[Any]) -> List[str]:
    sources = []
    for doc in docs:
        source = document_metadata(doc).get("source")
        if source:
            source_name = os.path.basename(str(source))
            if source_name not in sources:
                sources.append(source_name)
    return sources


def add_sources(answer: str, docs: List[Any]) -> str:
    sources = source_names(docs)
    if not sources or "\nSources:" in answer:
        return answer
    source_list = "\n".join(f"- {source}" for source in sources)
    return f"{answer.rstrip()}\n\nSources:\n{source_list}"

def build_prompt(
    context: str,
    question: str,
    conversation_history: List[Dict[str, str]] | None = None,
) -> str:
    q_lower = question.lower()
    history_text = ""
    if conversation_history:
        history_text = "\nConversation history:\n" + "\n".join(
            f"{turn['role'].title()}: {turn['content']}"
            for turn in conversation_history[-MAX_HISTORY_TURNS:]
        ) + "\n"
    if "leave" in q_lower:
        return f"""You are an assistant that answers questions using only the provided context.
    Use only facts explicitly supported by the context. Never guess, invent, or fill in missing values.
    If the answer is not supported by the context, reply exactly: "{NOT_FOUND_MESSAGE}"
    Provide a complete breakdown of employee leave (annual, sick, personal, totals) only when requested.

{history_text}
Context:
{context}

User question:
{question}

Answer concisely with all leave details. Include citations in square brackets referencing the source.
"""
    elif any(word in q_lower for word in POLICY_TERMS):
        return f"""You are an assistant that answers questions using only the provided context.
    Use only facts explicitly supported by the context. Never guess or invent consequences.
    If the answer is not supported by the context, reply exactly: "{NOT_FOUND_MESSAGE}"
Summarize the relevant rules, policies, or procedures from the context that apply to this situation. 
If the exact threshold for termination is not specified, explain that the policy uses progressive discipline (warnings, suspension, termination) rather than a fixed number of days.

{history_text}
Context:
{context}

User question:
{question}

Answer concisely. Include citations in square brackets referencing the source.
"""
    else:
        return f"""You are an assistant that answers questions using only the provided context.
    Use only facts explicitly supported by the context. Never guess or make unsupported claims.
    If the answer is not contained in the context, reply exactly: "{NOT_FOUND_MESSAGE}"

{history_text}
Context:
{context}

User question:
{question}

Answer concisely. Include citations in square brackets referencing the source.
"""

def call_llm(prompt: str) -> str:
    openai_key = os.getenv("OPENAI_API_KEY")
    if openai_key and OpenAI is not None:
        try:
            llm = OpenAI(model="gpt-4o-mini", temperature=0.0)
            return str(llm.invoke(prompt))
        except Exception as e:
            return f"[LLM call failed: {e}]"
    return "No LLM executed."

def rerank_results(query: str, docs: List[Any]) -> List[Any]:
    pairs = [(query, getattr(d, "page_content", None) or d.get("page_content", "")) for d in docs]
    scores = reranker.predict(pairs)
    scored_docs = list(zip(docs, scores))
    scored_docs.sort(key=lambda x: x[1], reverse=True)
    return [doc for doc, _ in scored_docs]

def format_employee_leave_record(record: Dict[str, Any], leave_category: str = "all") -> str:
    meta = record
    if leave_category == "annual":
        return f"""
{meta.get('Full Name','Unknown')} has taken {meta.get('Annual Leave Taken','N/A')} annual leave days.
Annual Leave Allocated: {meta.get('Annual Leave Allocated','N/A')}
Annual Leave Available: {meta.get('Annual Leave Available','N/A')}
""".strip()
    if leave_category == "sick":
        return f"""
{meta.get('Full Name','Unknown')} has taken {meta.get('Sick Leave Taken','N/A')} sick leave days.
Sick Leave Allocated: {meta.get('Sick Leave Allocated','N/A')}
Sick Leave Available: {meta.get('Sick Leave Available','N/A')}
""".strip()
    if leave_category == "personal":
        return f"""
{meta.get('Full Name','Unknown')} has taken {meta.get('Personal Leave Taken','N/A')} personal leave days.
Personal Leave Allocated: {meta.get('Personal Leave Allocated','N/A')}
Personal Leave Available: {meta.get('Personal Leave Available','N/A')}
""".strip()

    return f"""
Complete breakdown of employee leave for {meta.get('Full Name','Unknown')}:

- Annual Leave Allocated: {meta.get('Annual Leave Allocated','N/A')}
- Annual Leave Taken: {meta.get('Annual Leave Taken','N/A')}
- Annual Leave Available: {meta.get('Annual Leave Available','N/A')}
- Sick Leave Allocated: {meta.get('Sick Leave Allocated','N/A')}
- Sick Leave Taken: {meta.get('Sick Leave Taken','N/A')}
- Sick Leave Available: {meta.get('Sick Leave Available','N/A')}
- Personal Leave Allocated: {meta.get('Personal Leave Allocated','N/A')}
- Personal Leave Taken: {meta.get('Personal Leave Taken','N/A')}
- Personal Leave Available: {meta.get('Personal Leave Available','N/A')}
- Total Leaves Taken: {meta.get('Total Leaves Taken','N/A')}
- Total Leaves Available: {meta.get('Total Leaves Available','N/A')}
"""


def format_employee_record(record: Dict[str, Any]) -> str:
    return f"""Yes, {record.get('Full Name', 'This person')} is an employee of InnoCorp Solutions.

- Department: {record.get('Department', 'N/A')}
- Job Title: {record.get('Job Title', 'N/A')}
- Work Location: {record.get('Work Location', 'N/A')}
- Date of Joining: {record.get('Date of Joining', 'N/A')}
""".strip()

def load_employee_records() -> List[Dict[str, Any]]:
    if load_workbook is None:
        return []

    path = os.path.join(DATA_DIR, "InnoCorp_Solutions_Employee_Details.xlsx")
    if not os.path.exists(path):
        return []

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(header).strip() if header is not None else "" for header in next(rows)]
        records = [dict(zip(headers, row)) for row in rows]
        workbook.close()
        return [record for record in records if record.get("Full Name")]
    except Exception as exc:
        print(f"Could not read employee records: {exc}")
        return []


def answer_from_employee_data(
    query: str,
    docs: List[Any],
    employee_names: List[str],
    employee_records: List[Dict[str, Any]],
    current_question: str | None = None,
):
    query_lower = query.lower()
    intent_query_lower = (current_question or query).lower()

    current_matched_name = next(
        (
            name
            for name in employee_names
            if name.lower() in intent_query_lower
            or re.search(
                rf"\b{re.escape(name.split()[0].lower())}\b",
                intent_query_lower,
            )
        ),
        None,
    )
    has_employee_intent = is_employee_question(intent_query_lower)
    has_unrelated_intent = any(term in intent_query_lower for term in NON_EMPLOYEE_TERMS)
    if not has_employee_intent and not (current_matched_name and not has_unrelated_intent):
        return None

    if "annual" in intent_query_lower or "anual" in intent_query_lower:
        leave_category = "annual"
    elif "sick" in intent_query_lower:
        leave_category = "sick"
    elif "personal" in intent_query_lower:
        leave_category = "personal"
    else:
        leave_category = "all"

    matched_name = current_matched_name or next(
        (
            name
            for name in employee_names
            if name.lower() in query_lower
            or re.search(
                rf"\b{re.escape(name.split()[0].lower())}\b",
                query_lower,
            )
        ),
        None,
    )
    if not matched_name and re.search(r"\b(?:my|for|by|of)\s+[a-z]+(?:\s+[a-z]+)?\b", query_lower):
        return "No employee with that name was found in the employee database."

    if matched_name and employee_records:
        matched_record = next(
            (
                record
                for record in employee_records
                if str(record.get("Full Name", "")).lower() == matched_name.lower()
            ),
            None,
        )
        if matched_record:
            if is_leave_question(intent_query_lower):
                answer = format_employee_leave_record(matched_record, leave_category)
            else:
                answer = format_employee_record(matched_record)
            return add_sources(answer, [{"metadata": {"source": "InnoCorp_Solutions_Employee_Details.xlsx"}}])

    if matched_name:
        employee_docs = [doc for doc in docs if doc.metadata.get("Full Name","").lower() == matched_name.lower()]
        if employee_docs:
            if is_leave_question(intent_query_lower):
                answer = format_employee_leave_record(employee_docs[0].metadata, leave_category)
            else:
                answer = format_employee_record(employee_docs[0].metadata)
            return add_sources(answer, employee_docs[:1])
    return None


def build_search_query(question: str, conversation_history: List[Dict[str, str]]) -> str:
    """Resolve short follow-up questions using the recent conversation topic."""
    if not conversation_history:
        return question

    recent_user_questions = [
        turn["content"]
        for turn in conversation_history[-MAX_HISTORY_TURNS:]
        if turn.get("role") == "user"
    ]
    if not recent_user_questions:
        return question

    previous_questions = "\n".join(
        f"Previous question: {previous_question}"
        for previous_question in recent_user_questions
    )
    return f"{previous_questions}\nFollow-up question: {question}"

def main():
    print("Loading embeddings and Chroma DB...")
    emb = HuggingFaceEmbeddings(model_name=HF_MODEL_NAME)
    db = Chroma(persist_directory=CHROMA_DIR, embedding_function=emb)

    prov_map = load_provenance_with_fallback(PROVENANCE_FILE)
    print(f"Loaded provenance entries: {len(prov_map)}")

    stored_docs = db.get(include=["metadatas", "documents"])
    all_docs = stored_docs["documents"]
    all_metadatas = stored_docs.get("metadatas", [])
    bm25_documents = [
        Document(page_content=text, metadata=metadata or {})
        for text, metadata in zip(all_docs, all_metadatas)
    ]
    bm25_retriever = BM25Retriever.from_documents(bm25_documents)

    employee_records = load_employee_records()
    employee_names = [str(record["Full Name"]) for record in employee_records]
    conversation_history: List[Dict[str, str]] = []

    while True:
        q = input("\nEnter your question (or 'quit' to exit): ").strip()
        if not q:
            continue
        if q.lower() in {"quit", "exit"}:
            break

        search_query = build_search_query(q, conversation_history)
        current_query = q.lower()
        has_unrelated_intent = any(term in current_query for term in NON_EMPLOYEE_TERMS)
        is_employee_follow_up = not has_unrelated_intent and (
            is_employee_question(q) or any(
                is_employee_question(turn.get("content", ""))
                for turn in conversation_history
                if turn.get("role") == "user"
            )
        )
        vector_results = db.similarity_search(search_query, k=10)
        keyword_results = bm25_retriever.invoke(search_query)
        combined = vector_results + keyword_results
        reranked_results = rerank_results(search_query, combined)

        # Employee-specific check applies only to non-policy questions.
        answer = None
        if not any(term in q.lower() for term in POLICY_TERMS):
            answer = answer_from_employee_data(
                search_query,
                reranked_results,
                employee_names,
                employee_records,
                current_question=q,
            )
        if answer:
            print("\n" + answer.strip())
            conversation_history.extend([
                {"role": "user", "content": q},
                {"role": "assistant", "content": answer.strip()},
            ])
            conversation_history = conversation_history[-MAX_HISTORY_TURNS:]
            continue

        # If no local results, go external
        if not reranked_results and is_employee_follow_up:
            answer = NOT_FOUND_MESSAGE
            print("\n" + answer.strip())
            conversation_history.extend([
                {"role": "user", "content": q},
                {"role": "assistant", "content": answer.strip()},
            ])
            conversation_history = conversation_history[-MAX_HISTORY_TURNS:]
            continue

        if not reranked_results:
            answer = answer_from_external_search(q)
            print("\n" + answer.strip())
            conversation_history.extend([
                {"role": "user", "content": q},
                {"role": "assistant", "content": answer.strip()},
            ])
            conversation_history = conversation_history[-MAX_HISTORY_TURNS:]
            continue

        # Otherwise, use local context + LLM
        context_docs = reranked_results
        if any(term in q.lower() for term in POLICY_TERMS) or not is_employee_follow_up:
            context_docs = [
                doc
                for doc in reranked_results
                if "Employee_Details.xlsx" not in str(
                    (doc.metadata if hasattr(doc, "metadata") else doc.get("metadata", {})).get("source", "")
                )
            ]
        context = assemble_context(context_docs, MAX_CONTEXT_CHARS)
        prompt = build_prompt(context, search_query, conversation_history)
        answer = call_llm(prompt)

        if needs_external_search(answer):
            answer = NOT_FOUND_MESSAGE if is_employee_follow_up else answer_from_external_search(q)

        answer = add_sources(answer, context_docs)
        print("\n" + answer.strip())
        conversation_history.extend([
            {"role": "user", "content": q},
            {"role": "assistant", "content": answer.strip()},
        ])
        conversation_history = conversation_history[-MAX_HISTORY_TURNS:]

if __name__ == "__main__":
    main()
